"""
Excel Export Module for SnapGrade

This module provides functionality to export student grading data to Excel format.
"""

import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def convert_complex_to_str(value):
    """
    Convert a potentially complex value to a string representation
    
    Args:
        value: Any value that needs to be safely written to an Excel cell
        
    Returns:
        str: String representation of the value
    """
    if value is None:
        return ""
    elif isinstance(value, (str, int, float)):
        return str(value)
    elif isinstance(value, dict):
        # Check if this is a structured grading result
        if 'total_score' in value and 'questions' in value:
            return format_structured_grading_result(value)
        # Try to pretty format the dict
        try:
            return json.dumps(value, indent=2)
        except:
            result = ""
            for k, v in value.items():
                result += f"{k}: {convert_complex_to_str(v)}\n"
            return result
    elif isinstance(value, list):
        # Try to pretty format the list
        try:
            return json.dumps(value, indent=2)
        except:
            return "\n".join([convert_complex_to_str(item) for item in value])
    else:
        # Try general string representation
        return str(value)

def format_structured_grading_result(grading_data):
    """
    Format structured grading result into a readable string for Excel export
    
    Args:
        grading_data (dict): Structured grading result with total_score, questions, etc.
        
    Returns:
        str: Formatted string representation
    """
    try:
        formatted_parts = []
        
        # Overall score and percentage
        if 'total_score' in grading_data:
            formatted_parts.append(f"OVERALL SCORE: {grading_data['total_score']}")
        if 'percentage' in grading_data:
            formatted_parts.append(f"PERCENTAGE: {grading_data['percentage']}")
        
        formatted_parts.append("\n" + "="*50)
        
        # Question-by-question breakdown
        if 'questions' in grading_data and grading_data['questions']:
            formatted_parts.append("\nQUESTION BREAKDOWN:")
            for q in grading_data['questions']:
                q_num = q.get('question_number', '?')
                points_earned = q.get('points_earned', 0)
                points_possible = q.get('points_possible', 0)
                is_correct = q.get('is_correct', False)
                
                formatted_parts.append(f"\nQuestion {q_num}: {points_earned}/{points_possible} points {'✓' if is_correct else '✗'}")
                
                # Student answer
                if 'student_answer' in q:
                    student_ans = q['student_answer']
                    if isinstance(student_ans, dict):
                        formatted_parts.append("  Student Answer:")
                        for part, answer in student_ans.items():
                            formatted_parts.append(f"    {part}: {answer}")
                    else:
                        formatted_parts.append(f"  Student Answer: {student_ans}")
                
                # Correct answer
                if 'correct_answer' in q:
                    correct_ans = q['correct_answer']
                    if isinstance(correct_ans, dict):
                        formatted_parts.append("  Correct Answer:")
                        for part, answer in correct_ans.items():
                            formatted_parts.append(f"    {part}: {answer}")
                    else:
                        formatted_parts.append(f"  Correct Answer: {correct_ans}")
                
                # Grading rationale
                if 'grading_rationale' in q:
                    formatted_parts.append(f"  Rationale: {q['grading_rationale']}")
                
                # Teacher comment
                if 'teacher_comment' in q:
                    formatted_parts.append(f"  Teacher Comment: {q['teacher_comment']}")
                
                # OCR leniency note
                if q.get('ocr_leniency_applied'):
                    formatted_parts.append("  Note: OCR leniency applied")
        
        # Overall feedback section
        if 'overall_feedback' in grading_data:
            feedback = grading_data['overall_feedback']
            formatted_parts.append("\n" + "="*50)
            formatted_parts.append("\nOVERALL FEEDBACK:")
            
            if 'strengths' in feedback and feedback['strengths']:
                formatted_parts.append("\nStrengths:")
                for strength in feedback['strengths']:
                    formatted_parts.append(f"  • {strength}")
            
            if 'areas_for_improvement' in feedback and feedback['areas_for_improvement']:
                formatted_parts.append("\nAreas for Improvement:")
                for area in feedback['areas_for_improvement']:
                    formatted_parts.append(f"  • {area}")
            
            if 'next_steps' in feedback:
                formatted_parts.append(f"\nNext Steps: {feedback['next_steps']}")
        
        # Grading notes
        if 'grading_notes' in grading_data:
            formatted_parts.append(f"\nGrading Notes: {grading_data['grading_notes']}")
        
        return "\n".join(formatted_parts)
        
    except Exception as e:
        # Fallback to JSON if formatting fails
        try:
            return json.dumps(grading_data, indent=2)
        except:
            return str(grading_data)

def create_excel_for_batch_results(results, summary, assignment_name=None):
    """
    Create an Excel file from batch grading results using openpyxl
    
    Args:
        results (list): List of dictionaries containing grading results
        summary (dict): Summary information about the batch
        assignment_name (str, optional): Name of the assignment
        
    Returns:
        tuple: (file_path, filename) - Path to the saved Excel file and the filename
    """
    # Create a timestamp for the filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Generate a unique filename
    if assignment_name:
        safe_name = "".join([c if c.isalnum() else "_" for c in assignment_name])
        filename = f"{safe_name}_{timestamp}.xlsx"
    else:
        filename = f"batch_grading_results_{timestamp}.xlsx"
    
    # Create temp directory if it doesn't exist
    temp_dir = "temp_excel"
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
    
    file_path = os.path.join(temp_dir, filename)
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Create Summary sheet
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    
    # Add headers
    headers = ["Student Name", "Filename", "Score", "Status", "Feedback"]
    for col, header in enumerate(headers, 1):
        cell = summary_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    row = 2
    for result in results:
        try:
            result_filename = result.get("filename", "")
            student_name = result_filename.split('.')[0].replace('_', ' ')
            
            summary_sheet.cell(row=row, column=1).value = student_name
            summary_sheet.cell(row=row, column=2).value = result_filename
            summary_sheet.cell(row=row, column=3).value = convert_complex_to_str(result.get("score", "N/A"))
            
            if "error" in result:
                summary_sheet.cell(row=row, column=4).value = "Failed"
                summary_sheet.cell(row=row, column=5).value = convert_complex_to_str(result.get("error", "Unknown error"))
            else:
                summary_sheet.cell(row=row, column=4).value = "Processed"
                # Safely convert feedback to string
                summary_sheet.cell(row=row, column=5).value = convert_complex_to_str(result.get("feedback", ""))
            
            row += 1
        except Exception as e:
            print(f"Error adding row for {result.get('filename', 'unknown')}: {str(e)}")
            # Continue with next result
            continue
    
    # Add summary statistics at the bottom
    try:
        row += 2  # Add some space
        
        summary_sheet.cell(row=row, column=1).value = "Summary Statistics"
        summary_sheet.cell(row=row, column=1).font = Font(bold=True)
        
        stats = [
            ("Total Files", summary.get("total_files", 0)),
            ("Processed Files", summary.get("processed", 0)),
            ("Failed Files", summary.get("failed", 0)),
            ("Average Score", f"{summary.get('average_score', 0)}")
        ]
        
        row += 1
        for i, (label, value) in enumerate(stats):
            summary_sheet.cell(row=row + i, column=1).value = label
            summary_sheet.cell(row=row + i, column=2).value = convert_complex_to_str(value)
    except Exception as e:
        print(f"Error adding summary stats: {str(e)}")
    
    # Create Details sheet
    try:
        details_sheet = wb.create_sheet(title="Details")
        
        # Add headers for details sheet
        detail_headers = ["Student Name", "Filename", "Score", "Feedback", "Extracted Text"]
        for col, header in enumerate(detail_headers, 1):
            cell = details_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data to details sheet
        row = 2
        for result in results:
            if "error" not in result:
                try:
                    result_filename = result.get("filename", "")
                    student_name = result_filename.split('.')[0].replace('_', ' ')
                    
                    details_sheet.cell(row=row, column=1).value = student_name
                    details_sheet.cell(row=row, column=2).value = result_filename
                    details_sheet.cell(row=row, column=3).value = convert_complex_to_str(result.get("score", "N/A"))
                    
                    # Safely convert feedback and extracted text to strings
                    details_sheet.cell(row=row, column=4).value = convert_complex_to_str(result.get("feedback", ""))
                    
                    # Get extracted text, limited to avoid Excel cell size limits
                    extracted_text = convert_complex_to_str(result.get("extracted_text", ""))
                    if len(extracted_text) > 32000:  # Excel has a limit around 32,767 characters per cell
                        extracted_text = extracted_text[:32000] + "... (truncated)"
                    details_sheet.cell(row=row, column=5).value = extracted_text
                    
                    row += 1
                except Exception as e:
                    print(f"Error adding details row for {result.get('filename', 'unknown')}: {str(e)}")
                    # Continue with next result
                    continue
    except Exception as e:
        print(f"Error creating details sheet: {str(e)}")
    
    # Adjust column widths
    try:
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = min(cell_length, 100)  # Cap at 100 to avoid excessive widths
                        except:
                            pass
                
                adjusted_width = max_length + 2
                # Set a maximum width for text columns
                if column in ['E', 'D']:
                    adjusted_width = min(adjusted_width, 60)
                
                sheet.column_dimensions[column].width = adjusted_width
    except Exception as e:
        print(f"Error adjusting column widths: {str(e)}")
    
    # Save the workbook
    try:
        wb.save(file_path)
        print(f"Excel file saved successfully: {file_path}")
        return file_path, filename
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")
        # Try saving to a different location as a fallback
        fallback_path = os.path.join("static", filename)
        try:
            wb.save(fallback_path)
            print(f"Excel file saved to fallback location: {fallback_path}")
            return fallback_path, filename
        except Exception as fallback_error:
            print(f"Failed to save Excel file to fallback location: {str(fallback_error)}")
            raise


def save_excel_file(file_path, filename, directory="exports"):
    """
    Copy an Excel file to the specified directory if needed
    
    Args:
        file_path: Path to the source Excel file
        filename: Name for the destination file
        directory: Directory to save the file in
        
    Returns:
        str: Path to the saved file
    """
    # Check if source file exists
    if not os.path.exists(file_path):
        print(f"Warning: Source Excel file not found: {file_path}")
        return file_path
    
    # If the file is already in the right place, just return the path
    if os.path.dirname(file_path) == directory:
        return file_path
    
    # Create directory if it doesn't exist
    if not os.path.exists(directory):
        os.makedirs(directory)
    
    dest_path = os.path.join(directory, filename)
    
    # Copy the file
    import shutil
    try:
        shutil.copy2(file_path, dest_path)
        print(f"Excel file copied to: {dest_path}")
        
        # Verify the file was copied correctly
        if os.path.exists(dest_path) and os.path.getsize(dest_path) > 0:
            return dest_path
        else:
            print(f"Warning: Copied Excel file appears invalid at: {dest_path}")
            return file_path
    except Exception as e:
        print(f"Error copying Excel file: {str(e)}")
        return file_path
